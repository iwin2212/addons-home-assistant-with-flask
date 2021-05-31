import openpyxl
import json
from const import data_file
import os
# path to file climate excel
path = "1.xlsx"
climate_json_file = os.path.join(data_file, 'climate_code.json')


def create(mode_car: list, fan_mode_car: list, mode_dai: list, fan_mode_dai: list) -> dict:
    data = {}
    data["climate"] = {}
    return data


def add(data: dict, name: str, mode: list, fan_mode: list) -> dict:
    data["climate"][name] = {}
    data["climate"][name]["mode"] = mode
    data["climate"][name]["fan_mode"] = fan_mode
    return data


wb = openpyxl.load_workbook(path)
ws = wb.active
# columns = ["dry", "fan_only", "heat", "auto", "auto", "high", "medium", "low", "med", "max", "min", "Tên model điều hòa"]
index_colums = {}
start = 0
for col in range(1, 100):
    cell_obj = ws.cell(row=2, column=col)
    if cell_obj.value is not None:
        col_value = cell_obj.value
        index_colums[col] = col_value
name_tb = ""
mode = []
fan_mode = []
test = {}
test = create(test, None, None, None)
for row in range(3, 100):
    cell_name = ws.cell(row=row, column=2)
    if cell_name.value is not None:
        mode = []
        fan_mode = []
        for i in range(4, 10):
            if ws.cell(row=row, column=i).value == "x":
                mode.append(index_colums[i])
        for i in range(12, 19):
            if ws.cell(row=row, column=i).value == "x":
                fan_mode.append(index_colums[i])
        test = add(test, cell_name.value, mode, fan_mode)

with open(climate_json_file, 'w') as outfile:
    json.dump(test, outfile, indent=4)
